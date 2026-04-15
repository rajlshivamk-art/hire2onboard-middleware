from fastapi import APIRouter, HTTPException, status, UploadFile, File, BackgroundTasks, Depends, Query, Response, Request # Added Depends
from fastapi.responses import StreamingResponse, RedirectResponse
from typing import List, Optional
from datetime import datetime, timedelta, timezone, time
from beanie import PydanticObjectId
from ..models import Application, Feedback, OnboardingTask, User, Job, EmailTracking
from ..schemas import ApplicationCreate, ApplicationUpdate, FeedbackCreate, ApplicationResponse, FeedbackResponse, StageUpdate, TaskCreate, OfferCreate, TaskStatusUpdate, CandidateInteraction, UserResponse, EvaluationScore, OnboardingDocument
from ..utils.files import upload_file_from_stream, get_file_stream, get_gridfs
from ..utils.email_Fastmail import send_email
from .auth import get_current_user # Imported get_current_user
from ..services.erp_service import ERPService # Added ERPService
from fastapi import Form, Body
import traceback
import pandas as pd
import io
from io import BytesIO
import hashlib
import secrets
from bson import ObjectId
from zoneinfo import ZoneInfo 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from backend.config import settings
from urllib.parse import unquote, urlparse

router = APIRouter(
    prefix="/api/applications",
    tags=["applications"]
)

@router.get("/", response_model=List[ApplicationResponse])
async def get_applications(
    userId: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Company isolation (UNCHANGED)
    query = (
    Application.find(Application.companyId == current_user.companyId)
    if current_user.companyId
    else Application.find_all()
)

    if userId:
        if PydanticObjectId.is_valid(userId):
            if current_user.role == "Recruiter" and str(current_user.id) != userId:
                raise HTTPException(
                    status_code=403,
                    detail="Cannot view other recruiter's applications"
                )
            query = query.find(Application.assignedRecruiterId == userId)

    elif current_user.role == "Recruiter":
        query = query.find(Application.assignedRecruiterId == str(current_user.id))

    elif current_user.role == "Tech Interviewer":
        query = query.find(
        Application.interviewSchedules.interviewerId == str(current_user.id)
    )


    return await query.sort("-appliedDate").to_list()

@router.post("/", response_model=ApplicationResponse, status_code=status.HTTP_201_CREATED)
async def submit_application(application: ApplicationCreate, background_tasks: BackgroundTasks):
    app_data = application.model_dump()
    app_data["stage"] = "Applied"
    app_data["appliedDate"] = datetime.now()
    app_data["feedback"] = []
    app_data["onboardingTasks"] = []
    
    # --- Assign Company from Job ---
    if PydanticObjectId.is_valid(app_data["jobId"]):
        job = await Job.get(PydanticObjectId(app_data["jobId"]))
        if job:
            app_data["company"] = job.company
            app_data["companyId"] = job.companyId
            
    app_data["companyId"] = job.companyId if job else None
    app = Application(**app_data)
    await app.insert()

    # --- Email Triggers (Background) ---
    try:
        # Fetch Job Title for email context
        job_title = "Position"
        # Job already fetched above if valid
        if PydanticObjectId.is_valid(app.jobId):
             # We might have fetched it above, but ensure it's available
             if 'job' not in locals():
                 job = await Job.get(PydanticObjectId(app.jobId))
             
             if job:
                job_title = job.title

        # 1. Send specific confirmation to Candidate
        background_tasks.add_task(
            send_email,
            recipients=[app.email],
            subject=f"Application Received - {job_title}",
            template_name="candidate_application_confirmation",
            context={
        "name": app.name,
        "job_title": job_title,
        "application_id": str(app.id)   # <<< THIS IS REQUIRED
    }
        )
        
        # 2. Alert Recruiter
        background_tasks.add_task(
            send_email,
            recipients=["kunal.s@indianwellness.org"],
            subject=f"New Applicant Alert - {job_title}",
            template_name="recruiter_new_application_alert",
            context={
                "name": app.name, 
                "email": app.email,
                "job_title": job_title, 
            }
        )
        
        print("Scheduled background email tasks")
    except Exception as e:
        print(f"Error scheduling email tasks: {e}")
        print(f"Error scheduling email tasks: {e}")
    # ----------------------

    # --- ERP Sync (Applicant) ---
    background_tasks.add_task(ERPService.create_job_applicant, app_data, job_title)
    # ----------------------------

    return app

@router.get("/{applicationId}", response_model=ApplicationResponse)
async def get_application(
    applicationId: str,
    current_user: User = Depends(get_current_user)
):
    if not PydanticObjectId.is_valid(applicationId):
        raise HTTPException(status_code=404, detail="Application not found")
        
    application = await Application.get(PydanticObjectId(applicationId))
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    return application

@router.post("/{applicationId}/feedback", response_model=FeedbackResponse)
async def submit_feedback(applicationId: str, feedback: FeedbackCreate, background_tasks: BackgroundTasks):

    if application.companyId != current_user.companyId:
         raise HTTPException(status_code=403, detail="Unauthorized")

    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")
         
    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")
        
    fb_data = feedback.model_dump()
    fb_data["candidateId"] = applicationId
    fb_data["date"] = datetime.now()
    
    new_feedback = Feedback(**fb_data)
    app.feedback.append(new_feedback)
    await app.save()
    # --- ERP Sync (Feedback) ---
    job = await Job.get(PydanticObjectId(app.jobId))
    job_title = job.title if job else "Unknown Job"
    
    background_tasks.add_task(
         ERPService.sync_interview_feedback,
         feedback_data=fb_data,
         applicant_email=app.email,
         job_title=job_title
    )
    # ---------------------------
    
    return new_feedback

@router.put("/{applicationId}/feedback/{feedbackId}", response_model=FeedbackResponse)
async def update_feedback(applicationId: str, feedbackId: str, feedback: FeedbackCreate):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")
        
    found_fb = None
    for fb in app.feedback:
        if str(fb.id) == feedbackId:
            update_dict = feedback.model_dump()
            fb.roundName = update_dict["roundName"]
            fb.comments = update_dict["comments"]
            fb.decision = update_dict["decision"]
            fb.rating = update_dict.get("rating")
            fb.technicalSkills = update_dict.get("technicalSkills")
            fb.codeQuality = update_dict.get("codeQuality")
            fb.problemSolving = update_dict.get("problemSolving")
            fb.cultureFit = update_dict.get("cultureFit")
            fb.communication = update_dict.get("communication")
            fb.negotiatedSalary = update_dict.get("negotiatedSalary")
            found_fb = fb
            break
            
    if found_fb:
        await app.save()
        return found_fb
    else:
        raise HTTPException(status_code=404, detail="Feedback not found")

@router.delete("/{applicationId}/feedback/{feedbackId}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_feedback(applicationId: str, feedbackId: str):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    initial_len = len(app.feedback)
    app.feedback = [fb for fb in app.feedback if str(fb.id) != feedbackId]
    
    if len(app.feedback) < initial_len:
        await app.save()
    else:
        raise HTTPException(status_code=404, detail="Feedback not found")
    return None

@router.patch("/{applicationId}/stage", response_model=ApplicationResponse)
async def update_application_stage(applicationId: str, update: StageUpdate, background_tasks: BackgroundTasks):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
         raise HTTPException(status_code=404, detail="Application not found")

    app.stage = update.stage
    if update.reason:
        app.rejectionReason = update.reason
    
    if update.stage == "Hired":
        years = app.yearsOfExperience or 0
        
        if years < 1:
            required_docs = {
            "AADHAR",
            "PAN",
            "EDU_10TH",
            "EDU_12TH",
            "EDU_GRADUATION",
            "OFFER_ACCEPTANCE"
        }
            
        else:
            required_docs = {
            "AADHAR",
            "PAN",
            "EDU_10TH",
            "EDU_12TH",
            "EDU_GRADUATION",
            "EXPERIENCE",
            "OFFER_ACCEPTANCE"
        }
            
        uploaded_docs = {doc.type for doc in getattr(app, "documents", [])}
        missing = required_docs - uploaded_docs
        
        if missing:
            raise HTTPException(
            status_code=400,
            detail={
                "message": "Mandatory onboarding documents missing",
                "missingDocuments": list(missing),
                "candidateCategory": "FRESHER" if years < 1 else "EXPERIENCED"
            }
        )

    
    # Auto-generate onboarding tasks
    if update.stage == "Onboarding" and not app.onboardingTasks:
        default_tasks = [
            "Sign Offer Letter",
            "Background Check Completed",
            "IT Setup (Laptop/Email)",
            "HR Orientation",
            "Team Introduction",
            "Project Access Granted"
        ]
        app.onboardingTasks = [
            OnboardingTask(
                id=PydanticObjectId(),
                candidateId=applicationId,
                task=task_name,
                status="Pending"
            )
            for task_name in default_tasks
        ]
    
    await app.save()
    
    # --- Email Triggers (Background) ---
    try:
        # Fetch Job details
        job_title = "Position"
        if PydanticObjectId.is_valid(app.jobId):
            job = await Job.get(PydanticObjectId(app.jobId))
            if job:
                job_title = job.title
                
        # Custom message based on stage - ONLY for specific stages
        # User requested: Rejected, Offer, Onboarding (Document)
        should_send_email = False
        message = ""
        
        if update.stage == "Round 1":
            should_send_email = True
            message = "We are pleased to inform you that your application has been shortlisted for the next stage. Our team was impressed with your credentials and will be contacting you shortly to schedule an interaction."
        elif update.stage == "Rejected":
            should_send_email = True
            message = f"After a thorough review, we regret to inform you that we will not be proceeding with your application at this time. We had a strong pool of candidates and our decision was difficult. We wish you the best in your future endeavors."
        elif update.stage == "Offer":
            # For Offer, we NO LONGER send email automatically. 
            # It is triggered manually via /offer endpoint.
            should_send_email = False 
        elif update.stage == "Onboarding":
            # Don't send welcome email automatically. 
            # HR will trigger "Request Documents" manually or we send "Pending" reminders.
            should_send_email = False

        if should_send_email:
            background_tasks.add_task(
                send_email,
                recipients=[app.email],
                subject=f"Employment Offer - {job_title}" if update.stage == "Offer" else (f"Welcome Aboard - {job_title}" if update.stage == "Onboarding" else f"Application Update - {job_title}"),
                template_name="candidate_offer_letter" if update.stage == "Offer" else ("candidate_onboarding_welcome" if update.stage == "Onboarding" else "candidate_stage_update"),
                context={
                    "name": app.name, 
                    "job_title": job_title, 
                    "new_stage": update.stage,
                    "message": message,
                    "application_id": str(app.id),  
                    "documents": [
                        "Valid ID Proof (Aadhar/PAN)",
                        "Educational Certificates",
                        "Previous Employment Relieving Letter",
                        "Last 3 Months Salary Slips",
                        "Signed Offer Letter"
                    ] if update.stage == "Onboarding" else []
                }
            )
            print(f"Scheduled stage update email ({update.stage}) for {app.email}")
            print(f"Skipping email for stage: {update.stage}")

    except Exception as e:
        print(f"Error scheduling stage update email: {e}")
    # ----------------------
    
    # --- ERP Sync (Status Update) ---
    background_tasks.add_task(
        ERPService.update_applicant_status,
        applicant_email=app.email,
        status=update.stage
    )
    # --------------------------------
    
    return app

@router.post("/{applicationId}/offer", status_code=status.HTTP_202_ACCEPTED)
async def send_offer_email(
    applicationId: str, 
    background_tasks: BackgroundTasks,
    salary: float = Form(...),
    startDate: str = Form(...),
    additionalTerms: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None)
):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Fetch Job details
    job_title = "Position"
    if PydanticObjectId.is_valid(app.jobId):
        job = await Job.get(PydanticObjectId(app.jobId))
        if job:
            job_title = job.title

    attachments = [file] if file else None

    background_tasks.add_task(
        send_email,
        recipients=[app.email],
        subject=f"Employment Offer - {job_title}",
        template_name="candidate_offer_letter",
        context={
            "name": app.name,
            "job_title": job_title,
            "salary": salary,
            "start_date": startDate,
            "additional_terms": additionalTerms or "",
            "application_id": str(app.id),  
        },
        attachments=attachments
    )
    
    # Update candidate offer details (optional but good for record)
    app.offeredSalary = salary
    await app.save()

    # --- ERP Sync (Offer) ---
    background_tasks.add_task(
         ERPService.create_job_offer,
         offer_data={"salary": salary, "date": startDate}, 
         applicant_email=app.email, 
         job_title=job_title
    )
    # ------------------------

    return {"message": "Offer email scheduled"}

@router.patch("/{applicationId}/onboarding/{taskId}", response_model=ApplicationResponse)
async def update_onboarding_task(applicationId: str, taskId: str, status: TaskStatusUpdate):
    if not PydanticObjectId.is_valid(applicationId):
        raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    task_found = False
    for task in app.onboardingTasks:
        if str(task.id) == taskId:
            task.status = status.status
            if status.status == "Completed":
                task.completedDate = datetime.now()
            else:
                task.completedDate = None
            task_found = True
            break
    
    if not task_found:
        raise HTTPException(status_code=404, detail="Task not found")
        
    await app.save()
    return app

@router.post("/{applicationId}/onboarding/remind")
async def send_onboarding_reminder(applicationId: str, background_tasks: BackgroundTasks):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Find pending tasks (exclude Completed and Received)
    pending_tasks = [t.task for t in app.onboardingTasks if t.status == "Pending"]
    
    if not pending_tasks:
        return {"message": "No pending tasks to remind about (Received/Completed tasks excluded)"}

    # Fetch job title for subject
    job_title = "Position"
    if PydanticObjectId.is_valid(app.jobId):
        job = await Job.get(PydanticObjectId(app.jobId))
        if job:
            job_title = job.title

    background_tasks.add_task(
        send_email,
        recipients=[app.email],
        subject=f"Urgent: Pending Onboarding Documents - {job_title}",
        template_name="candidate_document_request",
        context={
            "name": app.name,
            "job_title": job_title,
            "documents": pending_tasks,
        }
    )
    
    return {"message": f"Reminder sent for {len(pending_tasks)} pending items"}

@router.patch("/{applicationId}/tasks/{taskId}/toggle", response_model=ApplicationResponse)
async def toggle_task(applicationId: str, taskId: str):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")
        
    found_task = False
    if app.onboardingTasks:
        for task in app.onboardingTasks:
            if str(task.id) == taskId:
                # Toggle between Pending and Completed
                if task.status == "Completed":
                    task.status = "Pending"
                    task.completedDate = None
                else:
                    task.status = "Completed"
                    task.completedDate = datetime.now()
                found_task = True
                break
                
    if found_task:
        await app.save()
        return app
    else:
        raise HTTPException(status_code=404, detail="Task not found")

@router.post("/{applicationId}/tasks", response_model=ApplicationResponse)
async def add_task(applicationId: str, task_in: TaskCreate):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    new_task = OnboardingTask(
        id=PydanticObjectId(),
        candidateId=applicationId,
        task=task_in.task,
        status="Pending"
    )
    if app.onboardingTasks is None:
        app.onboardingTasks = []
    app.onboardingTasks.append(new_task)
    
    await app.save()
    return app

@router.delete("/{applicationId}/tasks/{taskId}", response_model=ApplicationResponse)
async def delete_task(applicationId: str, taskId: str):
    if not PydanticObjectId.is_valid(applicationId):
         raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")
        
    if app.onboardingTasks:
        initial_len = len(app.onboardingTasks)
        app.onboardingTasks = [t for t in app.onboardingTasks if str(t.id) != taskId]
        if len(app.onboardingTasks) < initial_len:
            await app.save()
            return app
            
    raise HTTPException(status_code=404, detail="Task not found")

@router.patch("/{applicationId}", response_model=ApplicationResponse)
async def update_application(
    applicationId: str,
    app_update: ApplicationUpdate,
    current_user: User = Depends(get_current_user)
):
    if not PydanticObjectId.is_valid(applicationId):
        raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    update_data = app_update.model_dump(exclude_unset=True)

    # -------------------------------------------------
    # RBAC: Only HR / Manager / SuperAdmin can assign recruiter
    # -------------------------------------------------
    if "assignedRecruiterId" in update_data:
        if not (
            current_user.role in ["HR", "Manager", "SuperAdmin", "Admin"]
            or current_user.canManageUsers
        ):
            raise HTTPException(
                status_code=403,
                detail="You are not authorized to assign recruiter"
            )

    # -------------------------------------------------
    # STRICT MULTI-TENANCY CHECK (Job must belong to same company)
    # -------------------------------------------------
    if "jobId" in update_data and update_data["jobId"]:
        if not PydanticObjectId.is_valid(update_data["jobId"]):
            raise HTTPException(status_code=400, detail="Invalid jobId")

        job = await Job.get(PydanticObjectId(update_data["jobId"]))
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        if job.company != app.company:
            raise HTTPException(
                status_code=400,
                detail="Job does not belong to the same company"
            )

    # -------------------------------------------------
    # Apply updates safely
    # -------------------------------------------------
    for key, value in update_data.items():
        setattr(app, key, value)

    await app.save()
    return app

@router.post("/upload-resume", status_code=status.HTTP_201_CREATED)
async def upload_resume(file: UploadFile = File(...)):
    try:
        # Validate file type
        if file.content_type not in ["application/pdf", "application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
            raise HTTPException(status_code=400, detail="Invalid file type. Only PDF and Word documents are allowed.")
        
        file_id = await upload_file_from_stream(file.filename, file.file, file.content_type)
        
        # Return a URL-like string or just the ID that frontend can use to construct the URL
        # We'll return the full URL for the frontend to store in resumeUrl
        return {"url": f"/api/applications/resume/{file_id}", "fileId": file_id}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

@router.get("/resume/{file_id}")

async def get_resume(file_id: str):
    try:
        print(f"Attempting to download resume: {file_id}")
        grid_out = await get_file_stream(file_id)
        if not grid_out:
            print(f"Resume not found for ID: {file_id}")
            raise HTTPException(status_code=404, detail="Resume not found")
        
        # Define an async generator to yield chunks from the stream
        async def file_iterator():
            try:
                # Motor GridOut is an async iterator
                async for chunk in grid_out:
                    yield chunk
            except Exception as e:
                print(f"Error during file streaming: {e}")
                import traceback
                traceback.print_exc()
                raise

        # Properly quote the filename to handle special characters
        from urllib.parse import quote
        filename = quote(grid_out.filename)
        
        return StreamingResponse(
            file_iterator(), 
            media_type=grid_out.metadata.get("contentType", "application/octet-stream"),
            headers={"Content-Disposition": f"inline; filename*=utf-8''{filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in get_resume: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(e)}")

# Add a new candidate interaction
@router.post("/{application_id}/interactions", response_model=CandidateInteraction)
async def add_candidate_interaction(
    application_id: PydanticObjectId,
    interaction: CandidateInteraction,
    current_user = Depends(get_current_user)  # RBAC can be enforced inside this
):
    # Ensure only recruiter / HR assigned can add
    if current_user.role not in ["SuperAdmin", "HR", "Recruiter"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    app = await Application.get(application_id)
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    app.interactions.append(interaction)
    await app.save()
    return interaction

# Get all interactions for a candidate
@router.get("/{application_id}/interactions", response_model=List[CandidateInteraction])
async def get_candidate_interactions(
    application_id: PydanticObjectId,
    current_user = Depends(get_current_user)
):
    app = await Application.get(application_id)
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Only assigned recruiter, HR, or SuperAdmin can view
    if current_user.role not in ["SuperAdmin", "HR", "Recruiter"] and current_user.id != app.assignedRecruiterId:
        raise HTTPException(status_code=403, detail="Not authorized")

    return app.interactions

def recruiter_report_rbac(current_user: User = Depends(get_current_user)):
    allowed_roles = {"Manager", "Admin", "HR"}
    if current_user.role not in allowed_roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to access recruiter performance reports"
        )
    return current_user

BUSINESS_TZ = ZoneInfo("Asia/Kolkata")

def get_date_range(range_type: str, tz: timezone):
    today = datetime.now(tz).date()

    if range_type == "today":
        start_local = datetime.combine(today, time.min, tz)
        end_local = datetime.combine(today, time.max, tz)

    elif range_type == "weekly":
        start_local = datetime.combine(today - timedelta(days=6), time.min, tz)
        end_local = datetime.combine(today, time.max, tz)

    elif range_type == "monthly":
        start_local = datetime.combine(today.replace(day=1), time.min, tz)
        end_local = datetime.combine(today, time.max, tz)

    else:
        return None, None

    return (
        start_local.astimezone(timezone.utc),
        end_local.astimezone(timezone.utc)
    )

@router.get("/reports/recruiter-performance")
async def recruiter_performance_report(
    recruiter_id: Optional[PydanticObjectId] = Query(None, alias="recruiterId"),
    date_range: Optional[str] = Query(None, alias="dateRange", pattern="^(today|weekly|monthly)$"),
    start_date: Optional[datetime] = Query(None, alias="startDate"),
    end_date: Optional[datetime] = Query(None, alias="endDate"),
    current_user: User = Depends(recruiter_report_rbac),
):
    
    match_stage = {
        "companyId": current_user.companyId
    }

    if recruiter_id:
        match_stage["assignedRecruiterId"] = str(recruiter_id)

    if start_date and end_date:
        match_stage["appliedDate"] = {
            "$gte": start_date,
            "$lte": end_date
            }
    elif date_range:
        start, end = get_date_range(date_range, BUSINESS_TZ)
        match_stage["appliedDate"] = {
            "$gte": start,
            "$lte": end
            }    

    # =========================
    # KPI PIPELINE
    # =========================
    kpi_pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": None,
                "totalLineups": {"$sum": 1},
                "selected": {
                    "$sum": {"$cond": [{"$eq": ["$stage", "Hired"]}, 1, 0]}
                },
                "rejected": {
                    "$sum": {"$cond": [{"$eq": ["$stage", "Rejected"]}, 1, 0]}
                }
            }
        },
        {
            "$project": {
                "_id": 0,
                "totalLineups": 1,
                "selected": 1,
                "rejected": 1,
                "selectionRate": {
                    "$cond": [
                        {"$eq": ["$totalLineups", 0]},
                        0,
                        {
                            "$round": [
                                {
                                    "$multiply": [
                                        {"$divide": ["$selected", "$totalLineups"]},
                                        100
                                    ]
                                },
                                2
                            ]
                        }
                    ]
                }
            }
        }
    ]

    # =========================
    # RECRUITER-WISE PIPELINE
    # =========================
    pipeline = [
        {"$match": match_stage},
        {
            "$group": {
                "_id": "$assignedRecruiterId",
                "totalLineups": {"$sum": 1},
                "selected": {
                    "$sum": {"$cond": [{"$eq": ["$stage", "Hired"]}, 1, 0]}
                },
                "rejected": {
                    "$sum": {"$cond": [{"$eq": ["$stage", "Rejected"]}, 1, 0]}
                }
            }
        },
        {
            "$addFields": {
                "recruiterObjectId": {"$toObjectId": "$_id"}
            }
        },
        {
            "$lookup": {
                "from": "users",
                "localField": "recruiterObjectId",
                "foreignField": "_id",
                "as": "recruiter"
            }
        },
        {"$unwind": "$recruiter"},
        {
            "$project": {
                "recruiterId": "$_id",
                "recruiterName": "$recruiter.name",
                "totalLineups": 1,
                "selected": 1,
                "rejected": 1,
                "selectionRate": {
                    "$cond": [
                        {"$eq": ["$totalLineups", 0]},
                        0,
                        {
                            "$round": [
                                {
                                    "$multiply": [
                                        {"$divide": ["$selected", "$totalLineups"]},
                                        100
                                    ]
                                },
                                2
                            ]
                        }
                    ]
                }
            }
        },
        {"$sort": {"recruiterName": 1}}
    ]

    try:
        collection = Application.get_pymongo_collection()

        kpi_cursor = collection.aggregate(kpi_pipeline)
        kpi_data = await kpi_cursor.to_list(length=1)
        kpis = kpi_data[0] if kpi_data else {
            "totalLineups": 0,
            "selected": 0,
            "rejected": 0,
            "selectionRate": 0
        }

        cursor = collection.aggregate(pipeline)
        rows = await cursor.to_list(length=None)
        rows = rows or [] 

    except Exception:
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail="Failed to generate recruiter performance report"
        )

    return {
        "kpis": kpis,
        "rows": rows,
        "meta": {
            "company": current_user.company,
            "filters": {
                "recruiterId": str(recruiter_id) if recruiter_id else None,
                "dateRange": date_range,
                "startDate": start_date,
                "endDate": end_date
            }
        }
    }

@router.get("/reports/recruiter-performance/export")
async def recruiter_performance_export(
    recruiter_id: Optional[PydanticObjectId] = Query(None, alias="recruiterId"),
    date_range: Optional[str] = Query(None, alias="dateRange", pattern="^(today|weekly|monthly)$"),
    start_date: Optional[datetime] = Query(None, alias="startDate"),
    end_date: Optional[datetime] = Query(None, alias="endDate"),
    current_user: User = Depends(recruiter_report_rbac),
):
    try:
        match_stage = {"company": current_user.company}

        if recruiter_id:
            match_stage["assignedRecruiterId"] = str(recruiter_id)

        if start_date and end_date:
            match_stage["appliedDate"] = {
                "$gte": start_date,
                "$lte": end_date
            }
        elif date_range:
            start, end = get_date_range(date_range, BUSINESS_TZ)
            match_stage["appliedDate"] = {
                "$gte": start,
                "$lte": end
            }  

        kpi_pipeline = [
            {"$match": match_stage},
            {
                "$group": {
                    "_id": None,
                    "totalLineups": {"$sum": 1},
                    "selected": {"$sum": {"$cond": [{"$eq": ["$stage", "Hired"]}, 1, 0]}},
                    "rejected": {"$sum": {"$cond": [{"$eq": ["$stage", "Rejected"]}, 1, 0]}},
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "totalLineups": 1,
                    "selected": 1,
                    "rejected": 1,
                    "selectionRate": {
                        "$cond": [
                            {"$eq": ["$totalLineups", 0]},
                            0,
                            {"$round": [{"$multiply": [{"$divide": ["$selected", "$totalLineups"]}, 100]}, 2]},
                        ]
                    }
                }
            }
        ]

        pipeline = [
            {"$match": match_stage},
            {
                "$group": {
                    "_id": "$assignedRecruiterId",
                    "totalLineups": {"$sum": 1},
                    "selected": {"$sum": {"$cond": [{"$eq": ["$stage", "Hired"]}, 1, 0]}},
                    "rejected": {"$sum": {"$cond": [{"$eq": ["$stage", "Rejected"]}, 1, 0]}},
                }
            },
            {"$addFields": {"recruiterObjectId": {"$toObjectId": "$_id"}}},
            {
                "$lookup": {
                    "from": "users",
                    "localField": "recruiterObjectId",
                    "foreignField": "_id",
                    "as": "recruiter"
                }
            },
            {"$unwind": "$recruiter"},
            {
                "$project": {
                    "recruiterId": "$_id",
                    "recruiterName": "$recruiter.name",
                    "totalLineups": 1,
                    "selected": 1,
                    "rejected": 1,
                    "selectionRate": {
                        "$cond": [
                            {"$eq": ["$totalLineups", 0]},
                            0,
                            {"$round": [{"$multiply": [{"$divide": ["$selected", "$totalLineups"]}, 100]}, 2]}
                        ]
                    }
                }
            },
            {"$sort": {"recruiterName": 1}}
        ]

        collection = Application.get_pymongo_collection()

        kpi_cursor = collection.aggregate(kpi_pipeline)
        kpi_data = await kpi_cursor.to_list(length=1)
        kpis = kpi_data[0] if kpi_data else {
            "totalLineups": 0,
            "selected": 0,
            "rejected": 0,
            "selectionRate": 0
        }

        cursor = collection.aggregate(pipeline)
        rows = await cursor.to_list(length=None)
        rows = rows or [] 

        df_rows = pd.DataFrame(rows)
        df_kpis = pd.DataFrame([kpis])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_kpis.to_excel(writer, sheet_name="KPIs", index=False, startrow=4)
            df_rows.to_excel(writer, sheet_name="Recruiter Performance", index=False, startrow=4)

            wb = writer.book

            kpi_ws = wb["KPIs"]
            kpi_ws.merge_cells("A1:D1")
            kpi_ws["A1"] = "Recruiter Performance Report"
            kpi_ws["A1"].font = Font(size=16, bold=True)
            kpi_ws["A1"].alignment = Alignment(horizontal="center")

            kpi_ws.merge_cells("A2:D2")
            kpi_ws["A2"] = f"Company: {current_user.company} | Filters: {date_range or 'N/A'}"
            kpi_ws["A2"].alignment = Alignment(horizontal="center")

            for cell in kpi_ws[4]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFFF00")
                cell.alignment = Alignment(horizontal="center")

            for col in range(1, kpi_ws.max_column + 1):
                kpi_ws.column_dimensions[get_column_letter(col)].width = 22

            perf_ws = wb["Recruiter Performance"]
            perf_ws.merge_cells("A1:E1")
            perf_ws["A1"] = "Recruiter Performance Report"
            perf_ws["A1"].font = Font(size=16, bold=True)
            perf_ws["A1"].alignment = Alignment(horizontal="center")

            perf_ws.merge_cells("A2:E2")
            perf_ws["A2"] = f"Company: {current_user.company} | Filters: {date_range or 'N/A'}"
            perf_ws["A2"].alignment = Alignment(horizontal="center")

            for cell in perf_ws[4]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFFF00")
                cell.alignment = Alignment(horizontal="center")

            for col in range(1, perf_ws.max_column + 1):
                perf_ws.column_dimensions[get_column_letter(col)].width = 24

        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="recruiter_report.xlsx"'}
        )

    except Exception:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Failed to export report")
    
def recruiter_list_rbac(current_user: User = Depends(get_current_user)):
    allowed_roles = {"Manager", "Admin", "HR"}
    if current_user.role not in allowed_roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to access recruiter list"
        )
    return current_user

@router.get("/users", response_model=List[UserResponse])
async def get_recruiters(
    role: Optional[str] = Query(None),
    current_user: User = Depends(recruiter_list_rbac)
):
    if role != "Recruiter":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only role=Recruiter is supported for this endpoint"
        )

    recruiters = await User.find({
    "companyId": current_user.companyId,
    "role": "Recruiter"
    }).to_list()
    return recruiters

@router.post("/{applicationId}/evaluation")
async def add_evaluation_score(
    applicationId: str,
    score: EvaluationScore,
    current_user: User = Depends(get_current_user)
):
    # RBAC
    if current_user.role not in ["HR", "Manager", "SuperAdmin", "Admin", "Tech Interviewer", "Recruiter"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(404, "Application not found")

    # Company isolation
    if app.companyId and current_user.companyId and app.companyId != current_user.companyId:
        raise HTTPException(status_code=403, detail="Cross-company access denied")

    # **Validate roundId**
    if not score.roundId:
        raise HTTPException(status_code=400, detail="roundId is required")

    # Prevent duplicate round entry by same reviewer
    for s in app.evaluationScores:
        if s.roundId == score.roundId and s.reviewerId == score.reviewerId:
            raise HTTPException(
                status_code=400,
                detail="You already submitted evaluation for this round"
            )

    # --- ROUND LOGIC: Ensure 2 decimals ---
    score.technical = round(score.technical, 2) if score.technical is not None else None
    score.communication = round(score.communication, 2) if score.communication is not None else None
    score.problemSolving = round(score.problemSolving, 2) if score.problemSolving is not None else None
    score.cultureFit = round(score.cultureFit, 2) if score.cultureFit is not None else None

    # Calculate overall from fields if not provided
    if score.overall is None:
        fields = [v for v in [score.technical, score.communication, score.problemSolving, score.cultureFit] if v is not None]
        score.overall = round(sum(fields) / len(fields), 2) if fields else None
    else:
        score.overall = round(score.overall, 2)

    app.evaluationScores.append(score)

    # Cumulative logic (Industry standard) - rounded to 2 decimals
    valid_scores = [s.overall for s in app.evaluationScores if s.overall is not None]
    app.cumulativeScore = round(sum(valid_scores) / len(valid_scores), 2) if valid_scores else None

    await app.save()
    return {"message": "Score added", "cumulativeScore": app.cumulativeScore}

@router.get("/{applicationId}/evaluation")
async def get_evaluation(applicationId: str, current_user: User = Depends(get_current_user)):
    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(404, "Application not found")

    # Company isolation
    if app.companyId and current_user.companyId and app.companyId != current_user.companyId:
        raise HTTPException(status_code=403, detail="Cross-company access denied")

    return {
        "rounds": app.evaluationScores,
        "cumulativeScore": app.cumulativeScore
    }

@router.post("/{applicationId}/onboarding/upload-link")
async def generate_onboarding_upload_link(
    applicationId: str,
    background_tasks: BackgroundTasks
):
    if not PydanticObjectId.is_valid(applicationId):
        raise HTTPException(status_code=404, detail="Application not found")

    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Generate token
    raw_token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()

    # Store hash + expiry (48h)
    app.uploadTokenHash = token_hash
    app.uploadTokenExpiry = datetime.utcnow() + timedelta(hours=48)

    await app.save()
    is_experienced = bool(app.yearsOfExperience and app.yearsOfExperience > 0)

    upload_url = (
    f"{settings.FRONTEND_URL}/onboarding/upload"
    f"?uploadToken={raw_token}&isExperienced={str(is_experienced).lower()}"
)

    # Send email
    expires_at = datetime.utcnow() + timedelta(hours=48)

    background_tasks.add_task(
        send_email,
        recipients=[app.email],
        subject="Upload Required Documents",
        template_name="candidate_document_upload",
        context={
            "name": app.name,
            "upload_link": upload_url,              # matches template
            "expires_at": expires_at.strftime("%d %b %Y, %I:%M %p"),  # matches template
            "application_id": applicationId, 
    }
)
    return {
        "message": "Upload link generated and emailed to candidate",
        "expiresAt": app.uploadTokenExpiry
    }

@router.post("/onboarding/upload")
async def upload_onboarding_document(
    token: str = Form(...),
    documentType: str = Form(...),
    file: UploadFile = File(...)
):
    # Validate token
    token_hash = hashlib.sha256(token.encode()).hexdigest()

    application = await Application.find_one({
        "uploadTokenHash": token_hash,
        "uploadTokenExpiry": {"$gt": datetime.utcnow()}
    })

    if not application:
        raise HTTPException(status_code=401, detail="Invalid or expired upload link")

    if not file:
        raise HTTPException(status_code=400, detail="File is required")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file")

    # Hash file
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    # Upload to GridFS
    file_id = await upload_file_from_stream(
        filename=file.filename,
        source_stream=BytesIO(file_bytes),
        content_type=file.content_type
    )

    # Replace document by type
    application.documents = [
        doc for doc in application.documents if doc.type != documentType
    ]

    application.documents.append(
        {
            "type": documentType,
            "fileId": file_id,
            "fileHash": file_hash,
            "originalName": file.filename,
            "mimeType": file.content_type,
            "uploadedAt": datetime.utcnow()
        }
    )

    await application.save()

    return {
        "message": "Document uploaded successfully",
        "documentType": documentType
    }

@router.get("/{applicationId}/onboarding/document/{documentType}")
async def get_onboarding_document(
    applicationId: str,
    documentType: str,
    current_user: User = Depends(get_current_user)
):
    # Fetch application
    app = await Application.get(PydanticObjectId(applicationId))
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    # Authorization
    if app.companyId != current_user.companyId:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Find document
    document: Optional[OnboardingDocument] = next(
        (doc for doc in app.documents if doc.type == documentType),
        None
    )
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    # Get GridFS bucket
    fs = await get_gridfs()  # ✅ this returns the bucket
    if not fs:
        raise HTTPException(status_code=500, detail="GridFS not initialized")

    # Open download stream
    grid_out = await fs.open_download_stream(PydanticObjectId(document.fileId))
    if not grid_out:
        raise HTTPException(status_code=404, detail="File missing in storage")

    # Stream file to client
    return StreamingResponse(
        io.BytesIO(await grid_out.read()),
        media_type=document.mimeType or "application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={document.originalName}"}
    )

# ==========================================================
# Configuration
# ==========================================================

ALLOWED_REDIRECT_DOMAINS = [
    "localhost",
    "127.0.0.1",
    "hrms-3nbj.onrender.com",
]

BOT_KEYWORDS = ["bot", "spider", "crawl", "preview", "scanner"]
# ==========================================================
# Utility Functions
# ==========================================================

def get_client_ip(request: Request) -> str:
    """
    Extract real client IP (supports proxy/load balancer setups).
    """
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()

    if request.client:
        return request.client.host

    return ""


def hash_ip(ip: str) -> str:
    """
    Hash IP for compliance (GDPR safe).
    """
    if not ip:
        return ""
    return hashlib.sha256(ip.encode()).hexdigest()


def classify_source(user_agent: str):
    """
    Detect bots and proxy sources.
    """
    ua = user_agent.lower()

    is_bot = any(keyword in ua for keyword in BOT_KEYWORDS)
    is_gmail_proxy = "googleimageproxy" in ua

    if is_bot:
        return "bot", "low"

    if is_gmail_proxy:
        return "gmail_proxy", "medium"

    return "direct", "high"

def validate_redirect(url: str) -> bool:
    if not url:
        return False

    parsed = urlparse(url)

    # Allow relative paths like "/dashboard"
    if not parsed.netloc:
        return True

    if parsed.hostname and parsed.hostname in ALLOWED_REDIRECT_DOMAINS:
        return True

    return False

# ==========================================================
# Email Open Tracking (Pixel)
# ==========================================================

@router.get("/email/track/{tracking_id}")
async def track_email_open(tracking_id: str, request: Request):

    tracking = await EmailTracking.find_one(
        EmailTracking.trackingId == tracking_id
    )

    # Always return pixel (even if not found)
    pixel = (
        b'GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00'
        b'\xff\xff\xff!\xf9\x04\x01\x00\x00\x00\x00,\x00'
        b'\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;'
    )

    response = Response(content=pixel, media_type="image/gif")

    # Anti-cache headers
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, proxy-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Surrogate-Control"] = "no-store"
    response.headers["Content-Disposition"] = "inline; filename=track.gif"

    if not tracking:
        return response

    now = datetime.utcnow()
    user_agent = request.headers.get("user-agent", "")
    raw_ip = get_client_ip(request)
    ip_hash = hash_ip(raw_ip)

    source, confidence = classify_source(user_agent)

    # Prefetch detection (opened too quickly)
    if tracking.sentAt:
        diff_seconds = (now - tracking.sentAt).total_seconds()
        if diff_seconds < 5:
            confidence = "low"

    # Prevent rapid duplicate hits
    if tracking.lastOpenedAt:
        if (now - tracking.lastOpenedAt).total_seconds() < 3:
            return response

    # Update counters
    tracking.openCount += 1

    if not tracking.openedAt:
        tracking.openedAt = now

    tracking.lastOpenedAt = now
    tracking.lastOpenIP = ip_hash
    tracking.lastUserAgent = user_agent
    tracking.lastOpenSource = source
    tracking.openConfidence = confidence

    # Event log
    tracking.events.append({
        "type": "open",
        "timestamp": now,
        "ip": ip_hash,
        "userAgent": user_agent,
        "source": source,
        "confidence": confidence,
    })

    await tracking.save()

    return response

@router.head("/email/track/{tracking_id}")
async def track_email_open_head(tracking_id: str):
    """
    Some email clients send HEAD before GET.
    """
    return Response(status_code=200)
# ==========================================================
# Email Click Tracking
# ==========================================================

@router.get("/email/click/{tracking_id}")
async def track_email_click(
    tracking_id: str,
    request: Request,
    redirect: str
):
    tracking = await EmailTracking.find_one(
        EmailTracking.trackingId == tracking_id
    )

    if not tracking:
        raise HTTPException(status_code=404, detail="Tracking not found")

    decoded_redirect = unquote(redirect)

    # --------------------------------------------------------
    # Security: Prevent Open Redirect
    # --------------------------------------------------------
    if not validate_redirect(decoded_redirect):
        raise HTTPException(status_code=400, detail="Invalid redirect URL")

    now = datetime.utcnow()
    user_agent = request.headers.get("user-agent", "")
    raw_ip = get_client_ip(request)
    ip_hash = hash_ip(raw_ip)

    source, confidence = classify_source(user_agent)

    # --------------------------------------------------------
    # Prevent rapid duplicate clicks (2 sec protection)
    # --------------------------------------------------------
    if tracking.clickedAt:
        if (now - tracking.clickedAt).total_seconds() < 2:
            return RedirectResponse(url=decoded_redirect, status_code=302)

    # --------------------------------------------------------
    # CLICK LOGIC
    # --------------------------------------------------------
    tracking.clickCount += 1

    if not tracking.clickedAt:
        tracking.clickedAt = now

    # --------------------------------------------------------
    # CLICK IMPLIES OPEN (Enterprise logic)
    # --------------------------------------------------------
    if not tracking.openedAt:
        tracking.openedAt = now
        tracking.lastOpenedAt = now
        tracking.openCount = 1
        tracking.openConfidence = "high"

    tracking.lastOpenIP = ip_hash
    tracking.lastUserAgent = user_agent
    tracking.lastOpenSource = source

    # --------------------------------------------------------
    # Event Logging
    # --------------------------------------------------------
    tracking.events.append({
        "type": "click",
        "timestamp": now,
        "ip": ip_hash,
        "userAgent": user_agent,
        "source": source,
        "confidence": confidence,
    })

    await tracking.save()

    return RedirectResponse(url=decoded_redirect, status_code=302)
